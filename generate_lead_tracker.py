#!/usr/bin/env python3
"""
KanbanPro — Lead Tracker 28 Prospectos/Semana
3 hojas: Prospectos | Dashboard | Config
Brand: Teal #00e6c7 | Gold #FFB800 | Navy #1a1a2e
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from datetime import date, timedelta
import os

OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'kanbanpro-lead-tracker.xlsx')

# ─────────────────────────────────────────────
# BRAND COLORS (openpyxl: no #)
# ─────────────────────────────────────────────
NAVY     = "1a1a2e"
SURFACE  = "16213e"
CARD     = "0f3460"
TEAL     = "00e6c7"
TEAL_DIM = "00b8a0"
GOLD     = "FFB800"
WHITE    = "FFFFFF"
GRAY     = "9ca3af"
GRAY_LT  = "d1d5db"
GREEN    = "22c55e"
GREEN_DK = "166534"
RED      = "ef4444"
RED_DK   = "7f1d1d"
YELLOW   = "FBBF24"
YELLOW_DK= "78350f"
BLACK    = "000000"

# ─────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(color=WHITE, size=10, bold=False, italic=False, name="Arial"):
    return Font(name=name, color=color, size=size, bold=bold, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_all(color="D1D5DB"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom(color=TEAL_DIM):
    s = Side(style="thin", color=color)
    return Border(bottom=s)

def style_header_cell(cell, text, bg=NAVY, fg=TEAL, size=10, bold=True):
    cell.value = text
    cell.font = font(fg, size, bold)
    cell.fill = fill(bg)
    cell.alignment = align("center")
    cell.border = border_all(SURFACE)

def style_data_cell(cell, value=None, fmt=None, fg=GRAY_LT, bg=SURFACE, bold=False):
    if value is not None:
        cell.value = value
    cell.font = font(fg, 10, bold, name="Calibri")
    cell.fill = fill(bg)
    cell.alignment = align("left", wrap=True)
    cell.border = border_all("1e2d4a")
    if fmt:
        cell.number_format = fmt


wb = Workbook()

# ═══════════════════════════════════════════════════════
# SHEET 1 — CONFIG (hidden options for dropdowns)
# ═══════════════════════════════════════════════════════
cfg = wb.active
cfg.title = "Config"
cfg.sheet_state = "hidden"

cfg_lists = {
    "A": ("Tipo",      ["Gym", "Restaurante", "Clinica", "Spa", "Academia", "Otro"]),
    "B": ("Barrio",    ["Chapinero", "Usaquen", "Suba", "Engativa", "Kennnedy", "Santa Fe", "Teusaquillo", "Otro"]),
    "C": ("Estado",    ["Frio", "Tibio", "Caliente", "Compro", "Descartado", "Sin respuesta"]),
    "D": ("Material",  ["Nada", "One-Pager", "Deck", "One-Pager + Deck", "Solo mensaje"]),
    "E": ("Fuente",    ["Google Maps", "Referido", "Instagram", "LinkedIn", "Otro"]),
}

for col, (header, items) in cfg_lists.items():
    cfg[f"{col}1"] = header
    cfg[f"{col}1"].font = Font(bold=True)
    for row, item in enumerate(items, start=2):
        cfg[f"{col}{row}"] = item

# ═══════════════════════════════════════════════════════
# SHEET 2 — PROSPECTOS (core tracker)
# ═══════════════════════════════════════════════════════
ws = wb.create_sheet("Prospectos")
ws.sheet_view.showGridLines = False

# Tab color teal
ws.sheet_properties.tabColor = TEAL

# Freeze header + first col
ws.freeze_panes = "C2"

# ── TITLE ROW ──
ws.row_dimensions[1].height = 32
ws.merge_cells("A1:O1")
title_cell = ws["A1"]
title_cell.value = "KanbanPro — Tracker de Prospectos"
title_cell.font = font(TEAL, 14, True)
title_cell.fill = fill(NAVY)
title_cell.alignment = align("left", "center")

# ── COLUMN HEADERS (row 2) ──
headers = [
    ("#",               4),
    ("Negocio",         22),
    ("Tipo",            13),
    ("Barrio",          13),
    ("Contacto",        16),
    ("WhatsApp",        15),
    ("Fuente",          13),
    ("1er Contacto",    13),
    ("Material Env.",   13),
    ("Estado",          13),
    ("Ult. Contacto",   13),
    ("Prox. Follow-up", 15),
    ("Dias Restantes",  13),
    ("Notas",           28),
    ("Ingreso USD",     12),
]

ws.row_dimensions[2].height = 26
for col_idx, (hdr, width) in enumerate(headers, start=1):
    col_letter = get_column_letter(col_idx)
    cell = ws.cell(row=2, column=col_idx)
    style_header_cell(cell, hdr, NAVY, TEAL, 9, True)
    ws.column_dimensions[col_letter].width = width

# ── DATA ROWS (5 sample + 95 empty) ──
today = date.today()
sample_data = [
    (1, "CrossFit Norte", "Gym",          "Chapinero", "Carlos M.",  "321 555 0011", "Google Maps",
     today - timedelta(days=3), "One-Pager",        "Tibio",    today - timedelta(days=1), today + timedelta(days=1)),
    (2, "Iron Gym",       "Gym",          "Usaquen",   "Sandra R.",  "312 444 0099", "Google Maps",
     today - timedelta(days=5), "One-Pager + Deck", "Caliente", today - timedelta(days=2), today),
    (3, "FitLife Studio", "Gym",          "Suba",      "Diego L.",   "310 777 0033", "Google Maps",
     today - timedelta(days=7), "Nada",             "Frio",     today - timedelta(days=7), today + timedelta(days=3)),
    (4, "Muscle Zone",    "Gym",          "Chapinero", "Maria F.",   "315 888 0044", "Referido",
     today - timedelta(days=2), "Solo mensaje",     "Sin respuesta", today - timedelta(days=2), today + timedelta(days=2)),
    (5, "PowerHouse GYM", "Gym",          "Usaquen",   "Juan P.",    "318 999 0055", "Google Maps",
     today - timedelta(days=1), "One-Pager",        "Compro",   today,              today + timedelta(days=30)),
]

# Row color alternating
ROW_COLORS = [SURFACE, NAVY]

for row_offset, data in enumerate(sample_data):
    r = row_offset + 3
    bg = ROW_COLORS[row_offset % 2]
    ws.row_dimensions[r].height = 22

    num, negocio, tipo, barrio, contacto, wp, fuente, fecha1, material, estado, ult, prox = data

    values = [
        (num,      "0",        BLACK, False),
        (negocio,  "@",        WHITE, True ),
        (tipo,     "@",        TEAL,  False),
        (barrio,   "@",        GRAY_LT, False),
        (contacto, "@",        WHITE, False),
        (wp,       "@",        GRAY_LT, False),
        (fuente,   "@",        GRAY_LT, False),
        (fecha1,   "DD/MM/YY", TEAL,  False),
        (material, "@",        GRAY_LT, False),
        (estado,   "@",        GOLD,  True ),
        (ult,      "DD/MM/YY", GRAY_LT, False),
        (prox,     "DD/MM/YY", GOLD,  False),
    ]
    for col_idx, (val, fmt, fg, bld) in enumerate(values, start=1):
        cell = ws.cell(row=r, column=col_idx)
        style_data_cell(cell, val, fmt if fmt != "@" else None, fg, bg, bld)
        if fmt and fmt != "@" and isinstance(val, date):
            cell.number_format = "DD/MM/YY"

    # Col M — Dias Restantes (formula)
    cell_m = ws.cell(row=r, column=13)
    cell_m.value = f"=L{r}-TODAY()"
    cell_m.font = font(WHITE, 10, False, name="Calibri")
    cell_m.fill = fill(bg)
    cell_m.alignment = align("center")
    cell_m.border = border_all("1e2d4a")
    cell_m.number_format = '0" días"'

    # Col N — Notas (blank)
    cell_n = ws.cell(row=r, column=14)
    style_data_cell(cell_n, "", None, GRAY, bg)

    # Col O — Ingreso (formula: if Compro then 22 else 0)
    cell_o = ws.cell(row=r, column=15)
    cell_o.value = f'=IF(J{r}="Compro",22,0)'
    cell_o.font = font(GOLD if estado == "Compro" else GRAY_LT, 10, False, name="Calibri")
    cell_o.fill = fill(bg)
    cell_o.alignment = align("center")
    cell_o.border = border_all("1e2d4a")
    cell_o.number_format = '"$"#,##0.00'

# Empty rows 8–102
for r in range(8, 103):
    bg = ROW_COLORS[r % 2]
    ws.row_dimensions[r].height = 21
    num = r - 2
    # Auto-number
    cell_a = ws.cell(row=r, column=1)
    cell_a.value = num
    cell_a.font = font(GRAY, 9)
    cell_a.fill = fill(bg)
    cell_a.alignment = align("center")
    cell_a.border = border_all("1e2d4a")
    # Empty styled cells B–O
    for col_idx in range(2, 16):
        cell = ws.cell(row=r, column=col_idx)
        cell.fill = fill(bg)
        cell.border = border_all("1e2d4a")
        cell.font = font(GRAY_LT, 10, name="Calibri")
        if col_idx == 13:
            cell.value = f"=IF(L{r}=\"\",\"\",L{r}-TODAY())"
            cell.number_format = '0" días"'
            cell.alignment = align("center")
        elif col_idx == 15:
            cell.value = f'=IF(J{r}="Compro",22,0)'
            cell.number_format = '"$"#,##0.00'
            cell.alignment = align("center")

# ── DATA VALIDATION (dropdowns) ──
def make_dv(formula1, prompt):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.prompt = prompt
    dv.showErrorMessage = False
    return dv

dv_tipo     = make_dv('=Config!$A$2:$A$7',  "Selecciona tipo")
dv_barrio   = make_dv('=Config!$B$2:$B$9',  "Selecciona barrio")
dv_estado   = make_dv('=Config!$C$2:$C$7',  "Selecciona estado")
dv_material = make_dv('=Config!$D$2:$D$6',  "Selecciona material")
dv_fuente   = make_dv('=Config!$E$2:$E$6',  "Selecciona fuente")

ws.add_data_validation(dv_tipo)
ws.add_data_validation(dv_barrio)
ws.add_data_validation(dv_estado)
ws.add_data_validation(dv_material)
ws.add_data_validation(dv_fuente)

dv_tipo.add("C3:C102")
dv_barrio.add("D3:D102")
dv_fuente.add("G3:G102")
dv_material.add("I3:I102")
dv_estado.add("J3:J102")

# ── CONDITIONAL FORMATTING — Dias Restantes ──
# Rojo: <= 0 (atrasado)
ws.conditional_formatting.add(
    "M3:M102",
    FormulaRule(
        formula=['AND(M3<>"",M3<=0)'],
        fill=PatternFill("solid", fgColor=RED_DK),
        font=Font(color=RED, bold=True, name="Calibri", size=10),
    )
)
# Amarillo: 1-3 dias
ws.conditional_formatting.add(
    "M3:M102",
    FormulaRule(
        formula=['AND(M3>0,M3<=3)'],
        fill=PatternFill("solid", fgColor=YELLOW_DK),
        font=Font(color=YELLOW, bold=True, name="Calibri", size=10),
    )
)
# Verde: > 3 dias
ws.conditional_formatting.add(
    "M3:M102",
    FormulaRule(
        formula=['AND(M3>3)'],
        fill=PatternFill("solid", fgColor=GREEN_DK),
        font=Font(color=GREEN, name="Calibri", size=10),
    )
)

# ── CONDITIONAL FORMATTING — Estado column highlight ──
ws.conditional_formatting.add(
    "J3:J102",
    FormulaRule(
        formula=['J3="Compro"'],
        fill=PatternFill("solid", fgColor=GREEN_DK),
        font=Font(color=GREEN, bold=True, name="Calibri", size=10),
    )
)
ws.conditional_formatting.add(
    "J3:J102",
    FormulaRule(
        formula=['J3="Caliente"'],
        fill=PatternFill("solid", fgColor="7c2d12"),
        font=Font(color="fb923c", bold=True, name="Calibri", size=10),
    )
)

# ═══════════════════════════════════════════════════════
# SHEET 3 — DASHBOARD
# ═══════════════════════════════════════════════════════
dash = wb.create_sheet("Dashboard")
dash.sheet_view.showGridLines = False
dash.sheet_properties.tabColor = GOLD

# Title
dash.row_dimensions[1].height = 36
dash.merge_cells("A1:H1")
t = dash["A1"]
t.value = "KanbanPro — Dashboard de Ventas"
t.font = font(GOLD, 16, True)
t.fill = fill(NAVY)
t.alignment = align("left", "center")

dash.column_dimensions["A"].width = 26
dash.column_dimensions["B"].width = 16
dash.column_dimensions["C"].width = 4
dash.column_dimensions["D"].width = 22
dash.column_dimensions["E"].width = 16
dash.column_dimensions["F"].width = 4
dash.column_dimensions["G"].width = 22
dash.column_dimensions["H"].width = 16

# ── KPI SECTION ──
kpi_header_row = 3
dash.merge_cells(f"A{kpi_header_row}:H{kpi_header_row}")
kpi_h = dash[f"A{kpi_header_row}"]
kpi_h.value = "KPIs Generales"
kpi_h.font = font(TEAL, 11, True)
kpi_h.fill = fill(SURFACE)
kpi_h.alignment = align("left", "center")
dash.row_dimensions[kpi_header_row].height = 22

kpis = [
    ("Total contactados",           '=COUNTA(Prospectos!B3:B102)',                    GRAY_LT, '"0"'),
    ("Respondieron",                '=COUNTIF(Prospectos!J3:J102,"<>Frio")-COUNTIF(Prospectos!J3:J102,"Sin respuesta")', GRAY_LT, '"0"'),
    ("Calientes",                   '=COUNTIF(Prospectos!J3:J102,"Caliente")',          GOLD,    '"0"'),
    ("Compraron",                   '=COUNTIF(Prospectos!J3:J102,"Compro")',             GREEN,   '"0"'),
    ("Tasa conversion %",           '=IFERROR(COUNTIF(Prospectos!J3:J102,"Compro")/COUNTA(Prospectos!B3:B102),0)', GOLD, '0.0%'),
    ("Tasa respuesta %",            '=IFERROR((COUNTA(Prospectos!J3:J102)-COUNTIF(Prospectos!J3:J102,"Frio")-COUNTIF(Prospectos!J3:J102,"Sin respuesta"))/COUNTA(Prospectos!B3:B102),0)', TEAL, '0.0%'),
    ("Ingresos totales USD",        '=SUMIF(Prospectos!J3:J102,"Compro",Prospectos!O3:O102)', GREEN, '"$"#,##0.00'),
    ("Proyeccion mensual (x4.3 sem)",'=IFERROR(B9*4.3,0)',                              GOLD,   '"$"#,##0.00'),
]

for i, (label, formula, color, fmt) in enumerate(kpis):
    r = kpi_header_row + 1 + i
    dash.row_dimensions[r].height = 22

    col_a = dash.cell(row=r, column=1)
    col_a.value = label
    col_a.font = font(GRAY_LT, 10, name="Calibri")
    col_a.fill = fill(SURFACE)
    col_a.alignment = align("left", "center")
    col_a.border = border_bottom()

    col_b = dash.cell(row=r, column=2)
    col_b.value = formula
    col_b.font = Font(name="Calibri", color=color, size=13, bold=True)
    col_b.fill = fill(NAVY)
    col_b.alignment = align("center", "center")
    col_b.border = border_all(SURFACE)
    col_b.number_format = fmt

# ── POR TIPO DE NEGOCIO ──
tipo_start = kpi_header_row + len(kpis) + 3
dash.row_dimensions[tipo_start - 1].height = 22
dash.merge_cells(f"A{tipo_start-1}:B{tipo_start-1}")
th = dash.cell(row=tipo_start - 1, column=1)
th.value = "Conversion por Tipo de Negocio"
th.font = font(TEAL, 10, True)
th.fill = fill(SURFACE)
th.alignment = align("center")

tipos = ["Gym", "Restaurante", "Clinica", "Spa", "Academia", "Otro"]
style_header_cell(dash.cell(row=tipo_start, column=1), "Tipo",          NAVY, TEAL, 9)
style_header_cell(dash.cell(row=tipo_start, column=2), "Contactados",   NAVY, TEAL, 9)

for i, tipo in enumerate(tipos):
    r = tipo_start + 1 + i
    bg = ROW_COLORS[i % 2]
    cell_a = dash.cell(row=r, column=1)
    cell_a.value = tipo
    cell_a.font = font(GRAY_LT, 10, name="Calibri")
    cell_a.fill = fill(bg)
    cell_a.alignment = align("left", "center")
    cell_a.border = border_all("1e2d4a")

    cell_b = dash.cell(row=r, column=2)
    cell_b.value = f'=COUNTIF(Prospectos!C3:C102,A{r})'
    cell_b.font = Font(name="Calibri", color=TEAL, size=11, bold=True)
    cell_b.fill = fill(bg)
    cell_b.alignment = align("center")
    cell_b.border = border_all("1e2d4a")
    cell_b.number_format = "0"

# ── POR BARRIO ──
barrio_start = tipo_start
dash.merge_cells(f"D{barrio_start-1}:E{barrio_start-1}")
bh = dash.cell(row=barrio_start - 1, column=4)
bh.value = "Conversion por Barrio"
bh.font = font(TEAL, 10, True)
bh.fill = fill(SURFACE)
bh.alignment = align("center")

style_header_cell(dash.cell(row=barrio_start, column=4), "Barrio",       NAVY, TEAL, 9)
style_header_cell(dash.cell(row=barrio_start, column=5), "Contactados",  NAVY, TEAL, 9)

barrios = ["Chapinero", "Usaquen", "Suba", "Engativa", "Kennedy", "Santa Fe", "Teusaquillo", "Otro"]
for i, barrio in enumerate(barrios):
    r = barrio_start + 1 + i
    bg = ROW_COLORS[i % 2]
    cell_d = dash.cell(row=r, column=4)
    cell_d.value = barrio
    cell_d.font = font(GRAY_LT, 10, name="Calibri")
    cell_d.fill = fill(bg)
    cell_d.alignment = align("left", "center")
    cell_d.border = border_all("1e2d4a")

    cell_e = dash.cell(row=r, column=5)
    cell_e.value = f'=COUNTIF(Prospectos!D3:D102,D{r})'
    cell_e.font = Font(name="Calibri", color=GOLD, size=11, bold=True)
    cell_e.fill = fill(bg)
    cell_e.alignment = align("center")
    cell_e.border = border_all("1e2d4a")
    cell_e.number_format = "0"

# ── META SEMANAL ──
meta_start = tipo_start + len(tipos) + 3
dash.merge_cells(f"A{meta_start}:H{meta_start}")
meta_t = dash.cell(row=meta_start, column=1)
meta_t.value = "Meta Semanal — 28 contactos / semana"
meta_t.font = font(GOLD, 11, True)
meta_t.fill = fill(SURFACE)
meta_t.alignment = align("center")
dash.row_dimensions[meta_start].height = 22

meta_rows = [
    ("Contactos esta semana", '=COUNTIFS(Prospectos!H3:H102,">="&(TODAY()-WEEKDAY(TODAY(),2)+1),Prospectos!H3:H102,"<="&(TODAY()-WEEKDAY(TODAY(),2)+7))', TEAL),
    ("Meta semanal",          28,                                                                                                                           GRAY_LT),
    ("Cumplimiento %",        f'=IFERROR(B{meta_start+1}/B{meta_start+2},0)',                                                                              GOLD),
    ("Ventas necesarias/sem para 10 USD/mes", '=IFERROR(CEILING(10/22,1),"")',                                                                             GRAY_LT),
]

for i, (label, val, color) in enumerate(meta_rows):
    r = meta_start + 1 + i
    bg = ROW_COLORS[i % 2]
    cell_a = dash.cell(row=r, column=1)
    cell_a.value = label
    cell_a.font = font(GRAY_LT, 10, name="Calibri")
    cell_a.fill = fill(bg)
    cell_a.alignment = align("left", "center")
    cell_a.border = border_bottom()

    cell_b = dash.cell(row=r, column=2)
    cell_b.value = val
    cell_b.font = Font(name="Calibri", color=color, size=12, bold=True)
    cell_b.fill = fill(NAVY)
    cell_b.alignment = align("center")
    cell_b.border = border_all(SURFACE)
    if "%" in label:
        cell_b.number_format = "0.0%"

# Instrucciones de uso
inst_row = meta_start + len(meta_rows) + 3
dash.merge_cells(f"A{inst_row}:H{inst_row}")
inst_t = dash.cell(row=inst_row, column=1)
inst_t.value = "USO: Actualiza la hoja Prospectos. Este dashboard recalcula automaticamente."
inst_t.font = font(TEAL_DIM, 9, False, True)
inst_t.fill = fill(NAVY)
inst_t.alignment = align("center")

# ─────────────────────────────────────────────
# GUARDAR
# ─────────────────────────────────────────────
wb.save(OUTPUT)
print(f"[OK] Tracker generado: {OUTPUT}")
